#!/usr/bin/env python3
"""
Matomo ARK Extractor v2.1
Extraction des statistiques ARK depuis les exports Matomo XML
avec récupération des métadonnées via l'API OAI-PMH du catalogue Portfolio

Bibliothèques spécialisées de la Ville de Paris
"""

import os
import sys
import re
import threading
import xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
from pathlib import Path
import webbrowser
import urllib.parse
import platform
import subprocess
import shutil

# Interface moderne
import customtkinter as ctk
from tkinter import filedialog, messagebox

# Traitement données
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Détection du système pour choisir la méthode HTTP
IS_WINDOWS = platform.system() == 'Windows'
IS_MACOS = platform.system() == 'Darwin'

# Sur Windows, utiliser requests ; sur macOS/Linux, utiliser curl
if IS_WINDOWS:
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Désactiver les warnings
import warnings
warnings.filterwarnings('ignore')

# Configuration CustomTkinter
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Couleurs personnalisées
COLORS = {
    'primary': '#1f538d',
    'secondary': '#14375e', 
    'accent': '#3a7ebf',
    'success': '#2d8659',
    'warning': '#c9a227',
    'error': '#bf3636',
    'bg_dark': '#1a1a2e',
    'bg_card': '#16213e',
    'text': '#e8e8e8',
    'text_muted': '#a0a0a0'
}

# Configuration OAI-PMH
OAI_BASE_URL = "https://bibliotheques-specialisees.paris.fr/in/rest/oai"
OAI_IDENTIFIER_PREFIX = "oai:bibliotheques-specialisees.paris.fr:"


class MatomoARKExtractor(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Configuration fenêtre
        self.title("📚 Matomo ARK Extractor v2.1 - Bibliothèques spécialisées Paris")
        self.geometry("1100x900")
        self.minsize(950, 750)
        
        # Variables
        self.xml_path = ctk.StringVar()
        self.status_text = ctk.StringVar(value="Sélectionnez un fichier XML Matomo")
        self.progress_value = ctk.DoubleVar(value=0)
        self.scrape_metadata = ctk.BooleanVar(value=True)
        self.include_components = ctk.BooleanVar(value=False)
        self.ark_data = []
        self.is_processing = False
        
        # Interface
        self.create_ui()
        
    def create_ui(self):
        # Frame principal avec padding
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        self.create_header()
        
        # Zone de sélection fichier
        self.create_file_selector()
        
        # Options
        self.create_options()
        
        # Boutons d'action
        self.create_action_buttons()
        
        # Barre de progression
        self.create_progress_section()
        
        # Zone de résultats/aperçu
        self.create_results_section()
        
        # Footer
        self.create_footer()
    
    def create_header(self):
        header_frame = ctk.CTkFrame(self.main_frame, fg_color=COLORS['bg_card'], corner_radius=15)
        header_frame.pack(fill="x", pady=(0, 15))
        
        # Titre avec icône
        title_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_frame.pack(pady=20, padx=20)
        
        ctk.CTkLabel(
            title_frame, 
            text="📚 Matomo ARK Extractor",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=COLORS['text']
        ).pack()
        
        ctk.CTkLabel(
            title_frame,
            text="Extraction des statistiques + métadonnées via API OAI-PMH",
            font=ctk.CTkFont(size=14),
            text_color=COLORS['text_muted']
        ).pack(pady=(5, 0))
        
        # Badges info
        badges_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        badges_frame.pack(pady=(0, 15))
        
        for text, color in [("Bibliothèques spécialisées", COLORS['primary']), 
                            ("Ville de Paris", COLORS['accent']),
                            ("v2.1 - OAI-PMH", COLORS['success'])]:
            badge = ctk.CTkLabel(
                badges_frame,
                text=text,
                font=ctk.CTkFont(size=11),
                fg_color=color,
                corner_radius=12,
                padx=12,
                pady=4
            )
            badge.pack(side="left", padx=5)
    
    def create_file_selector(self):
        file_frame = ctk.CTkFrame(self.main_frame, fg_color=COLORS['bg_card'], corner_radius=15)
        file_frame.pack(fill="x", pady=(0, 15))
        
        inner_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        inner_frame.pack(fill="x", padx=20, pady=20)
        
        ctk.CTkLabel(
            inner_frame,
            text="📁 Fichier XML Matomo",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS['text']
        ).pack(anchor="w")
        
        # Ligne de sélection
        select_frame = ctk.CTkFrame(inner_frame, fg_color="transparent")
        select_frame.pack(fill="x", pady=(10, 0))
        
        self.file_entry = ctk.CTkEntry(
            select_frame,
            textvariable=self.xml_path,
            placeholder_text="Cliquez sur 'Parcourir' pour sélectionner le fichier XML...",
            font=ctk.CTkFont(size=13),
            height=45,
            corner_radius=10
        )
        self.file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        self.browse_btn = ctk.CTkButton(
            select_frame,
            text="📂 Parcourir",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=45,
            width=140,
            corner_radius=10,
            fg_color=COLORS['primary'],
            hover_color=COLORS['accent'],
            command=self.browse_file
        )
        self.browse_btn.pack(side="right")
    
    def create_options(self):
        options_frame = ctk.CTkFrame(self.main_frame, fg_color=COLORS['bg_card'], corner_radius=15)
        options_frame.pack(fill="x", pady=(0, 15))
        
        inner_frame = ctk.CTkFrame(options_frame, fg_color="transparent")
        inner_frame.pack(fill="x", padx=20, pady=15)
        
        ctk.CTkLabel(
            inner_frame,
            text="⚙️ Options",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS['text']
        ).pack(anchor="w")
        
        # Checkbox pour métadonnées OAI-PMH
        self.metadata_check = ctk.CTkCheckBox(
            inner_frame,
            text="Récupérer les métadonnées via API OAI-PMH (titre, auteur, date, type...)",
            variable=self.scrape_metadata,
            font=ctk.CTkFont(size=13),
            checkbox_height=22,
            checkbox_width=22,
            corner_radius=5
        )
        self.metadata_check.pack(anchor="w", pady=(10, 0))
        
        ctk.CTkLabel(
            inner_frame,
            text="✅ Utilise l'API OAI-PMH du catalogue (décocher si réseau bloqué)",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['success']
        ).pack(anchor="w", padx=(28, 0), pady=(2, 0))
        
        # Checkbox pour composantes
        self.components_check = ctk.CTkCheckBox(
            inner_frame,
            text="Inclure les composantes/vues (BAP..., pages numérisées)",
            variable=self.include_components,
            font=ctk.CTkFont(size=13),
            checkbox_height=22,
            checkbox_width=22,
            corner_radius=5
        )
        self.components_check.pack(anchor="w", pady=(10, 0))
        
        ctk.CTkLabel(
            inner_frame,
            text="📄 Crée une feuille supplémentaire avec le détail par composante",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['text_muted']
        ).pack(anchor="w", padx=(28, 0), pady=(2, 0))
    
    def create_action_buttons(self):
        btn_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 15))
        
        # Bouton principal
        self.run_btn = ctk.CTkButton(
            btn_frame,
            text="▶️  Extraire et générer l'Excel",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=55,
            corner_radius=12,
            fg_color=COLORS['success'],
            hover_color="#238c4d",
            command=self.start_extraction
        )
        self.run_btn.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        # Bouton aperçu
        self.preview_btn = ctk.CTkButton(
            btn_frame,
            text="👁️ Aperçu",
            font=ctk.CTkFont(size=14),
            height=55,
            width=120,
            corner_radius=12,
            fg_color=COLORS['secondary'],
            hover_color=COLORS['primary'],
            command=self.show_preview
        )
        self.preview_btn.pack(side="right")
    
    def create_progress_section(self):
        self.progress_frame = ctk.CTkFrame(self.main_frame, fg_color=COLORS['bg_card'], corner_radius=15)
        self.progress_frame.pack(fill="x", pady=(0, 15))
        
        inner_frame = ctk.CTkFrame(self.progress_frame, fg_color="transparent")
        inner_frame.pack(fill="x", padx=20, pady=15)
        
        # Status
        self.status_label = ctk.CTkLabel(
            inner_frame,
            textvariable=self.status_text,
            font=ctk.CTkFont(size=13),
            text_color=COLORS['text']
        )
        self.status_label.pack(anchor="w")
        
        # Barre de progression
        self.progress_bar = ctk.CTkProgressBar(
            inner_frame,
            variable=self.progress_value,
            height=12,
            corner_radius=6,
            progress_color=COLORS['accent']
        )
        self.progress_bar.pack(fill="x", pady=(10, 0))
        self.progress_bar.set(0)
    
    def create_results_section(self):
        results_frame = ctk.CTkFrame(self.main_frame, fg_color=COLORS['bg_card'], corner_radius=15)
        results_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        inner_frame = ctk.CTkFrame(results_frame, fg_color="transparent")
        inner_frame.pack(fill="both", expand=True, padx=20, pady=15)
        
        # Titre section
        header_row = ctk.CTkFrame(inner_frame, fg_color="transparent")
        header_row.pack(fill="x")
        
        ctk.CTkLabel(
            header_row,
            text="📊 Journal",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLORS['text']
        ).pack(side="left")
        
        self.count_label = ctk.CTkLabel(
            header_row,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=COLORS['text_muted']
        )
        self.count_label.pack(side="right")
        
        # Zone de texte pour les logs
        self.log_textbox = ctk.CTkTextbox(
            inner_frame,
            font=ctk.CTkFont(family="Consolas", size=12),
            corner_radius=10,
            fg_color="#0d1117",
            text_color="#c9d1d9",
            height=250
        )
        self.log_textbox.pack(fill="both", expand=True, pady=(10, 0))
        self.log("🚀 Prêt ! Sélectionnez un fichier XML Matomo pour commencer.")
        self.log("")
        self.log("ℹ️  Cette version utilise l'API OAI-PMH pour récupérer les métadonnées.")
        self.log("   Endpoint: " + OAI_BASE_URL)
    
    def create_footer(self):
        footer_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        footer_frame.pack(fill="x")
        
        ctk.CTkLabel(
            footer_frame,
            text="CCPID - Bibliothèques de la Ville de Paris",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['text_muted']
        ).pack(side="left")
        
        ctk.CTkLabel(
            footer_frame,
            text="v2.1 - OAI-PMH",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['accent']
        ).pack(side="right")
    
    def log(self, message, level="INFO"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        icons = {"INFO": "ℹ️", "SUCCESS": "✅", "ERROR": "❌", "WARNING": "⚠️", "PROGRESS": "🔄", "DATA": "📄"}
        icon = icons.get(level, "")
        
        if level == "INFO" and message.startswith("ℹ️"):
            # Déjà formaté
            self.log_textbox.insert("end", f"{message}\n")
        elif level == "INFO" and message == "":
            self.log_textbox.insert("end", "\n")
        else:
            self.log_textbox.insert("end", f"[{timestamp}] {icon} {message}\n")
        
        self.log_textbox.see("end")
        self.update_idletasks()
    
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Sélectionner le fichier XML Matomo",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
        )
        if filename:
            self.xml_path.set(filename)
            self.log(f"Fichier sélectionné: {Path(filename).name}", "SUCCESS")
            self.status_text.set(f"Fichier: {Path(filename).name}")
    
    def start_extraction(self):
        if not self.xml_path.get():
            messagebox.showwarning("Attention", "Veuillez sélectionner un fichier XML")
            return
        
        if not os.path.exists(self.xml_path.get()):
            messagebox.showerror("Erreur", "Le fichier sélectionné n'existe pas")
            return
        
        if self.is_processing:
            return
        
        self.is_processing = True
        self.run_btn.configure(state="disabled", text="⏳ Traitement en cours...")
        self.browse_btn.configure(state="disabled")
        self.progress_bar.set(0)
        self.log_textbox.delete("1.0", "end")
        
        # Lancer dans un thread
        thread = threading.Thread(target=self.extraction_thread, daemon=True)
        thread.start()
    
    def extraction_thread(self):
        try:
            self.log("Démarrage de l'extraction...", "PROGRESS")
            
            # 1. Parser le XML
            self.status_text.set("Analyse du fichier XML...")
            self.progress_value.set(0.1)
            self.ark_data, self.components_data = self.parse_xml(self.xml_path.get())
            
            if not self.ark_data:
                self.log("Aucune donnée ARK trouvée dans le fichier", "ERROR")
                return
            
            self.log(f"Trouvé {len(self.ark_data)} notices ARK uniques", "SUCCESS")
            if self.components_data:
                self.log(f"Trouvé {len(self.components_data)} composantes/vues", "SUCCESS")
            self.count_label.configure(text=f"{len(self.ark_data)} notices")
            
            # 2. Récupérer les métadonnées si demandé
            if self.scrape_metadata.get():
                self.status_text.set("Récupération des métadonnées via OAI-PMH...")
                self.progress_value.set(0.2)
                self.fetch_oai_metadata()
            
            # 3. Générer l'Excel
            self.status_text.set("Génération du fichier Excel...")
            self.progress_value.set(0.9)
            output_path = self.generate_excel()
            
            self.progress_value.set(1.0)
            self.status_text.set("Terminé !")
            self.log(f"Fichier Excel généré: {Path(output_path).name}", "SUCCESS")
            
            # Message de succès
            messagebox.showinfo(
                "Extraction terminée",
                f"Le fichier Excel a été créé:\n\n{output_path}"
            )
            
            # Ouvrir le dossier (compatible Windows et macOS)
            import subprocess
            import platform
            folder = os.path.dirname(output_path)
            if platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', folder])
            elif platform.system() == 'Windows':
                os.startfile(folder)
            else:  # Linux
                subprocess.call(['xdg-open', folder])
            
        except Exception as e:
            self.log(f"Erreur: {str(e)}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            messagebox.showerror("Erreur", f"Une erreur s'est produite:\n{str(e)}")
        
        finally:
            self.is_processing = False
            self.run_btn.configure(state="normal", text="▶️  Extraire et générer l'Excel")
            self.browse_btn.configure(state="normal")
    
    def parse_xml(self, xml_path):
        """Parse le fichier XML Matomo et extrait les données ARK"""
        self.log("Parsing du fichier XML...")
        
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        notices = []  # Niveau notice (pf..., FRCGM...)
        components = []  # Niveau composante (BAP..., vues...)
        
        def get_type_from_ark(ark_id):
            """Détermine le type de ressource depuis l'identifiant ARK"""
            if ark_id.startswith('FRCGMNOV'):
                return 'Fonds iconographique - Nouvelles'
            elif ark_id.startswith('FRCGMSUP'):
                return 'Fonds iconographique - Suppléments'
            elif ark_id.startswith('FRCGM'):
                return 'Fonds iconographique'
            elif ark_id.startswith('pf'):
                return 'Notice bibliographique'
            else:
                return 'Autre'
        
        def extract_rows(element, parent_ark=None):
            for row in element.findall('.//row'):
                label = row.findtext('label', '')
                url_elem = row.find('url')
                url = url_elem.text if url_elem is not None else None
                segment = row.findtext('segment', '')
                
                # Données Matomo
                data = {
                    'nb_visits': int(row.findtext('nb_visits', '0') or 0),
                    'nb_uniq_visitors': row.findtext('nb_uniq_visitors', '') or row.findtext('sum_daily_nb_uniq_visitors', ''),
                    'nb_hits': int(row.findtext('nb_hits', '0') or 0),
                    'sum_time_spent': int(row.findtext('sum_time_spent', '0') or 0),
                    'avg_time_on_page': row.findtext('avg_time_on_page', ''),
                    'bounce_rate': row.findtext('bounce_rate', ''),
                    'exit_rate': row.findtext('exit_rate', ''),
                    'entry_nb_visits': row.findtext('entry_nb_visits', ''),
                    'entry_bounce_count': row.findtext('entry_bounce_count', ''),
                    'exit_nb_visits': row.findtext('exit_nb_visits', ''),
                }
                
                # CAS 1: URL explicite avec ARK
                if url and '/ark:/' in url:
                    # Extraire l'ARK de l'URL - regex améliorée pour capturer tous les formats
                    ark_match = re.search(r'ark:/(\d+)/([a-zA-Z0-9\-_\.]+)(?:/([a-zA-Z0-9\-_\.]+))?', url)
                    if ark_match:
                        naan = ark_match.group(1)
                        ark_id_raw = ark_match.group(2)
                        component_id = ark_match.group(3)  # Peut être None
                        
                        # Nettoyer l'ark_id des suffixes comme .locale=fr ou .locale
                        ark_id = re.sub(r'\.locale(=.*)?$', '', ark_id_raw)
                        
                        ark_full = f"ark:/{naan}/{ark_id}"
                        
                        # Est-ce une composante ?
                        is_component = False
                        if component_id:
                            if (component_id.startswith('BAP') or 
                                component_id.startswith('BHP') or
                                component_id.startswith('BHD') or
                                re.match(r'^\d{4}$', component_id) or
                                component_id.startswith('A') or  # A2194500 etc
                                component_id.startswith('B')):   # B1454607 etc
                                is_component = True
                        
                        if is_component:
                            components.append({
                                'ark_notice': ark_full,
                                'component_id': component_id,
                                'url': url,
                                **data
                            })
                            # AUSSI ajouter la notice parente (sera agrégée/dédoublonnée plus tard)
                            clean_url = f"https://bibliotheques-specialisees.paris.fr/{ark_full}"
                            notices.append({
                                'ark': ark_full,
                                'ark_id': ark_id,
                                'naan': naan,
                                'url': clean_url,
                                'type': get_type_from_ark(ark_id),
                                'titre': '', 'auteur': '', 'contributeur': '',
                                'date': '', 'editeur': '', 'description': '',
                                'bibliotheque': '', 'cote': '', 'type_oai': '',
                                'sujet': '', 'format_doc': '', 'langue': '',
                                'droits': '', 'relation': '',
                                **data  # Les stats de la composante contribuent à la notice parente
                            })
                        else:
                            # C'est une notice - ON NE FILTRE PLUS les vues v0001/selectedTab
                            # On agrège ensuite par ARK donc les doublons ne posent pas problème
                            clean_url = re.sub(r'/v\d+\..*$', '', url)
                            clean_url = re.sub(r'\?.*$', '', clean_url)
                            
                            notices.append({
                                'ark': ark_full,
                                'ark_id': ark_id,
                                'naan': naan,
                                'url': clean_url,
                                'type': get_type_from_ark(ark_id),
                                'titre': '', 'auteur': '', 'contributeur': '',
                                'date': '', 'editeur': '', 'description': '',
                                'bibliotheque': '', 'cote': '', 'type_oai': '',
                                'sujet': '', 'format_doc': '', 'langue': '',
                                'droits': '', 'relation': '',
                                **data
                            })
                
                # CAS 1bis: Pas d'URL mais ARK encodé dans le segment
                elif not url and segment and 'ark%253A%252F' in segment:
                    seg_match = re.search(r'ark%253A%252F(\d+)%252F([a-zA-Z0-9\-]+)', segment)
                    if seg_match:
                        naan = seg_match.group(1)
                        ark_id = seg_match.group(2)
                        ark_full = f"ark:/{naan}/{ark_id}"
                        
                        notices.append({
                            'ark': ark_full,
                            'ark_id': ark_id,
                            'naan': naan,
                            'url': f"https://bibliotheques-specialisees.paris.fr/ark:/{naan}/{ark_id}",
                            'type': get_type_from_ark(ark_id),
                            'titre': '', 'auteur': '', 'contributeur': '',
                            'date': '', 'editeur': '', 'description': '',
                            'bibliotheque': '', 'cote': '', 'type_oai': '',
                            'sujet': '', 'format_doc': '', 'langue': '',
                            'droits': '', 'relation': '',
                            **data
                        })
                
                # CAS 2: Label qui est un identifiant de notice (niveau 3)
                elif label and not label.startswith('/') and label not in ['ark:', '73873', 'Autres']:
                    if re.match(r'^(pf|FRCGM)', label):
                        # Nettoyer le label des suffixes comme .locale=fr ou .locale
                        clean_label = re.sub(r'\.locale(=.*)?$', '', label)
                        ark_full = f"ark:/73873/{clean_label}"
                        notices.append({
                            'ark': ark_full,
                            'ark_id': clean_label,
                            'naan': '73873',
                            'url': f"https://bibliotheques-specialisees.paris.fr/ark:/73873/{clean_label}",
                            'type': get_type_from_ark(clean_label),
                            'titre': '', 'auteur': '', 'contributeur': '',
                            'date': '', 'editeur': '', 'description': '',
                            'bibliotheque': '', 'cote': '', 'type_oai': '',
                            'sujet': '', 'format_doc': '', 'langue': '',
                            'droits': '', 'relation': '',
                            **data
                        })
                
                # CAS 3: Label qui est une composante (/BAP..., /BHP..., /0001...)
                elif label and label.startswith('/'):
                    comp_id = label[1:]  # Enlever le /
                    if (comp_id.startswith('BAP') or 
                        comp_id.startswith('BHP') or 
                        comp_id.startswith('BHD') or
                        re.match(r'^\d{4}$', comp_id)):
                        # Essayer de reconstruire l'ARK parent depuis le segment
                        parent_match = re.search(r'ark%253A%252F(\d+)%252F([a-zA-Z0-9\-]+)', segment)
                        if parent_match:
                            parent_ark = f"ark:/{parent_match.group(1)}/{parent_match.group(2)}"
                        else:
                            parent_ark = "ark:/73873/inconnu"
                        
                        components.append({
                            'ark_notice': parent_ark,
                            'component_id': comp_id,
                            'url': url or '',
                            **data
                        })
        
        extract_rows(root)
        
        # Agréger par ARK unique (notices)
        aggregated = defaultdict(lambda: {
            'visits': 0, 'hits': 0, 'sum_time': 0,
            'uniq_visitors': 0,  # On prend le max car on ne peut pas additionner les visiteurs uniques
            'entry_visits': 0, 'entry_bounces': 0, 'exit_visits': 0,
            'data': None
        })
        
        for item in notices:
            ark = item['ark']
            aggregated[ark]['visits'] += item['nb_visits']
            aggregated[ark]['hits'] += item['nb_hits']
            aggregated[ark]['sum_time'] += item['sum_time_spent']
            # Visiteurs uniques: prendre le max (on ne peut pas les additionner)
            try:
                current_uniq = int(item.get('nb_uniq_visitors') or 0)
                if current_uniq > aggregated[ark]['uniq_visitors']:
                    aggregated[ark]['uniq_visitors'] = current_uniq
            except:
                pass
            try:
                aggregated[ark]['entry_visits'] += int(item.get('entry_nb_visits') or 0)
                aggregated[ark]['entry_bounces'] += int(item.get('entry_bounce_count') or 0)
                aggregated[ark]['exit_visits'] += int(item.get('exit_nb_visits') or 0)
            except:
                pass
            if aggregated[ark]['data'] is None:
                aggregated[ark]['data'] = item
        
        # Construire la liste finale des notices
        result_notices = []
        for ark, agg in aggregated.items():
            data = agg['data'].copy()
            data['nb_visits'] = agg['visits']
            data['nb_hits'] = agg['hits']
            data['sum_time_spent'] = agg['sum_time']
            data['nb_uniq_visitors'] = agg['uniq_visitors'] if agg['uniq_visitors'] > 0 else ''
            data['entry_nb_visits'] = agg['entry_visits']
            data['entry_bounce_count'] = agg['entry_bounces']
            data['exit_nb_visits'] = agg['exit_visits']
            result_notices.append(data)
        
        # Trier par visites
        result_notices.sort(key=lambda x: x['nb_visits'], reverse=True)
        
        # Logger les top 5
        self.log("Top 5 des notices les plus consultées:")
        for i, item in enumerate(result_notices[:5], 1):
            self.log(f"  #{i}: {item['ark_id']} - {item['nb_visits']} visites", "DATA")
        
        return result_notices, components
    
    def fetch_oai_metadata(self):
        """Récupère les métadonnées via l'API OAI-PMH - teste plusieurs formats"""
        total = len(self.ark_data)
        self.log(f"Récupération des métadonnées pour {total} notices via OAI-PMH...")
        self.log(f"Endpoint: {OAI_BASE_URL}")
        self.log(f"Préfixe OAI: {OAI_IDENTIFIER_PREFIX}")
        
        # Formats à tester dans l'ordre de priorité
        METADATA_PREFIXES = ["oai_dc_syracuse", "oai_dc", "inmedia"]
        self.log(f"Formats testés: {', '.join(METADATA_PREFIXES)}")
        
        success_count = 0
        error_count = 0
        no_record_count = 0
        
        import time
        
        for i, item in enumerate(self.ark_data):
            # Délai entre les requêtes
            if i > 0:
                time.sleep(0.3)
            
            # Mise à jour progression
            progress = 0.2 + (i / max(total, 1)) * 0.6
            self.progress_value.set(progress)
            self.status_text.set(f"Métadonnées: {i+1}/{total} - {item['ark_id'][:20]}...")
            
            ark_identifier = item['ark']
            oai_identifier = f"{OAI_IDENTIFIER_PREFIX}{ark_identifier}"
            
            metadata = None
            last_response_text = None
            working_format = None
            
            # Tester chaque format jusqu'à en trouver un qui fonctionne
            for meta_prefix in METADATA_PREFIXES:
                oai_url = f"{OAI_BASE_URL}?verb=GetRecord&identifier={oai_identifier}&metadataPrefix={meta_prefix}"
                
                # Log détaillé pour les 3 premières notices
                if i < 3:
                    self.log(f"  Test {meta_prefix} pour {item['ark_id']}", "PROGRESS")
                
                try:
                    # Essayer curl d'abord (disponible sur Windows 10+, macOS, Linux)
                    curl_success = False
                    try:
                        # Trouver curl
                        curl_cmd = shutil.which('curl')
                        if not curl_cmd and IS_WINDOWS:
                            # Chemin par défaut sur Windows
                            curl_cmd = r'C:\Windows\System32\curl.exe'
                        
                        if curl_cmd:
                            result = subprocess.run(
                                [curl_cmd, '-s', '--max-time', '30', '-k', '--http1.0', oai_url],
                                capture_output=True,
                                text=True,
                                timeout=35
                            )
                            if result.returncode == 0 and result.stdout and '<' in result.stdout:
                                last_response_text = result.stdout
                                curl_success = True
                    except Exception:
                        pass
                    
                    if not curl_success:
                        # Fallback: urllib avec HTTP/1.0 pour éviter chunked encoding
                        import urllib.request
                        import ssl
                        import http.client
                        
                        # Forcer HTTP/1.0 (pas de chunked encoding)
                        http.client.HTTPConnection._http_vsn = 10
                        http.client.HTTPConnection._http_vsn_str = 'HTTP/1.0'
                        http.client.HTTPSConnection._http_vsn = 10
                        http.client.HTTPSConnection._http_vsn_str = 'HTTP/1.0'
                        
                        ssl_ctx = ssl.create_default_context()
                        ssl_ctx.check_hostname = False
                        ssl_ctx.verify_mode = ssl.CERT_NONE
                        
                        req = urllib.request.Request(oai_url)
                        req.add_header('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)')
                        req.add_header('Accept', 'application/xml')
                        req.add_header('Connection', 'close')
                        
                        with urllib.request.urlopen(req, timeout=30, context=ssl_ctx) as response:
                            last_response_text = response.read().decode('utf-8')
                        
                        # Restaurer HTTP/1.1 pour les autres requêtes
                        http.client.HTTPConnection._http_vsn = 11
                        http.client.HTTPConnection._http_vsn_str = 'HTTP/1.1'
                        http.client.HTTPSConnection._http_vsn = 11
                        http.client.HTTPSConnection._http_vsn_str = 'HTTP/1.1'
                    
                    if i < 3:
                        self.log(f"    → {len(last_response_text)} chars", "PROGRESS")
                    
                    # Vérifier les erreurs OAI
                    if 'idDoesNotExist' in last_response_text or 'noRecordsMatch' in last_response_text:
                        continue
                    
                    if '<error' in last_response_text and 'cannotDisseminateFormat' in last_response_text:
                        continue
                    
                    if '<error' in last_response_text:
                        continue
                    
                    # Parser la réponse
                    metadata = self.parse_oai_response(last_response_text)
                    
                    if metadata and metadata.get('title'):
                        working_format = meta_prefix
                        break  # On a trouvé un format qui fonctionne !
                        
                except Exception as e:
                    if i < 3:
                        self.log(f"    Exception: {str(e)[:50]}", "WARNING")
                    continue
            
            # Stocker les métadonnées si on en a trouvé
            if metadata and metadata.get('title'):
                item['titre'] = metadata.get('title', '')
                item['auteur'] = metadata.get('creator', '')
                item['contributeur'] = metadata.get('contributor', '')
                item['date'] = metadata.get('date', '')
                item['editeur'] = metadata.get('publisher', '')
                item['description'] = metadata.get('description', '')[:300] if metadata.get('description') else ''
                item['type_oai'] = metadata.get('type', '')
                item['sujet'] = metadata.get('subject', '')
                item['cote'] = metadata.get('identifier', '')
                item['bibliotheque'] = metadata.get('source', '')
                item['format_doc'] = metadata.get('format', '')
                item['langue'] = metadata.get('language', '')
                item['droits'] = metadata.get('rights', '')
                item['relation'] = metadata.get('relation', '')
                
                success_count += 1
                if success_count <= 5:
                    self.log(f"  ✓ [{working_format}] {item['ark_id']}: {item['titre'][:50]}...", "DATA")
            else:
                # Analyser pourquoi ça n'a pas marché
                if last_response_text:
                    if 'idDoesNotExist' in last_response_text or 'noRecordsMatch' in last_response_text:
                        no_record_count += 1
                        if no_record_count <= 3:
                            self.log(f"  Notice non trouvée: {item['ark_id']}", "WARNING")
                    else:
                        error_count += 1
                        if error_count <= 3:
                            self.log(f"  Pas de métadonnées pour {item['ark_id']}", "WARNING")
                else:
                    error_count += 1
        
        self.log(f"", "INFO")
        self.log(f"=== Bilan OAI-PMH ===", "INFO")
        self.log(f"Titres récupérés: {success_count} / {total}", "SUCCESS" if success_count > 0 else "WARNING")
        if no_record_count > 0:
            self.log(f"Non trouvés dans OAI: {no_record_count}", "WARNING")
        if error_count > 0:
            self.log(f"Erreurs/Sans métadonnées: {error_count}", "WARNING")
            if error_count > total * 0.5:
                self.log(f"", "INFO")
                self.log(f"💡 Beaucoup d'erreurs réseau ? Décochez 'Récupérer les métadonnées'", "INFO")
                self.log(f"   pour générer l'Excel sans titres (stats Matomo uniquement).", "INFO")
        
        # Enrichir les composantes avec le titre de leur notice parente
        if self.components_data:
            # Créer un dictionnaire ark (complet) → titre
            titles_map = {item['ark']: item.get('titre', '') for item in self.ark_data if item.get('titre')}
            
            comp_enriched = 0
            for comp in self.components_data:
                ark_notice = comp.get('ark_notice', '')
                if ark_notice and ark_notice in titles_map:
                    comp['titre_notice'] = titles_map[ark_notice]
                    comp_enriched += 1
            
            if comp_enriched > 0:
                self.log(f"Composantes enrichies avec titre notice parente: {comp_enriched}", "SUCCESS")
    
    def parse_oai_response(self, xml_text):
        """Parse la réponse XML OAI-PMH pour extraire les métadonnées (Dublin Core + inmedia)"""
        try:
            root = ET.fromstring(xml_text)
            metadata = {}
            
            # Liste complète des champs Dublin Core
            dc_fields = ['title', 'creator', 'date', 'publisher', 'description', 'type', 
                        'subject', 'identifier', 'source', 'format', 'rights', 'language', 
                        'relation', 'coverage', 'contributor']
            
            # 1. Chercher les champs Dublin Core (dc:title, dc:creator, etc.)
            for dc_elem in dc_fields:
                found_values = []
                
                for e in root.iter():
                    tag_local = e.tag.split('}')[-1] if '}' in e.tag else e.tag
                    if tag_local.lower() == dc_elem.lower() and e.text:
                        text = e.text.strip()
                        if text and text not in found_values:
                            found_values.append(text)
                
                if found_values:
                    if dc_elem in ['subject', 'type', 'rights']:
                        metadata[dc_elem] = ' | '.join(found_values)
                    elif dc_elem == 'identifier':
                        non_url = [v for v in found_values if not v.startswith('http') and not v.startswith('oai:')]
                        metadata[dc_elem] = ' | '.join(non_url) if non_url else ''
                    else:
                        metadata[dc_elem] = found_values[0]
            
            # 2. Compléter avec les propriétés inmedia si présentes
            # Format: <inmedia:property name="title">valeur</inmedia:property>
            for e in root.iter():
                tag_local = e.tag.split('}')[-1] if '}' in e.tag else e.tag
                if tag_local == 'property' and e.text:
                    name = (e.attrib.get('name') or '').lower()
                    value = e.text.strip()
                    if not value:
                        continue
                    
                    # Mapper les propriétés inmedia vers Dublin Core
                    if name == 'title' and not metadata.get('title'):
                        metadata['title'] = value
                    elif name in ('creator', 'author') and not metadata.get('creator'):
                        metadata['creator'] = value
                    elif name == 'date' and not metadata.get('date'):
                        metadata['date'] = value
                    elif name == 'publisher' and not metadata.get('publisher'):
                        metadata['publisher'] = value
                    elif name == 'description' and not metadata.get('description'):
                        metadata['description'] = value
                    elif name == 'subject':
                        prev = metadata.get('subject', '')
                        metadata['subject'] = (prev + ' | ' if prev else '') + value
                    elif name == 'source' and not metadata.get('source'):
                        metadata['source'] = value
                    elif name == 'ark' and not metadata.get('identifier'):
                        metadata['identifier'] = value
            
            return metadata if metadata else None
            
        except Exception as e:
            return None
    
    def generate_excel(self):
        """Génère le fichier Excel avec toutes les données"""
        self.log("Génération du fichier Excel...")
        
        # Chemin de sortie horodaté
        xml_dir = os.path.dirname(self.xml_path.get())
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"stats_matomo_ark_{timestamp}.xlsx"
        output_path = os.path.join(xml_dir, output_filename)
        
        wb = Workbook()
        
        # === Feuille 1: Données principales ===
        ws = wb.active
        ws.title = "Statistiques ARK"
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1f538d')
        alt_fill = PatternFill('solid', fgColor='e8f0fe')
        success_fill = PatternFill('solid', fgColor='d4edda')
        border = Border(
            left=Side(style='thin', color='cccccc'),
            right=Side(style='thin', color='cccccc'),
            top=Side(style='thin', color='cccccc'),
            bottom=Side(style='thin', color='cccccc')
        )
        link_font = Font(color='0563C1', underline='single')
        
        # En-têtes - Stats Matomo + TOUS les champs Dublin Core
        headers = [
            'Rang', 'ARK complet', 'ID ARK', 'Type ressource',
            # Métadonnées OAI-PMH
            'Titre', 'Auteur', 'Contributeur', 'Date', 'Éditeur', 
            'Bibliothèque / Source', 'Cote / Identifiant', 'Type document', 
            'Sujets', 'Format', 'Langue', 'Droits', 'Description',
            # Statistiques Matomo
            'Visites', 'Visiteurs uniques', 'Pages vues', 
            'Temps total (s)', 'Temps moyen', 'Taux rebond', 'Taux sortie',
            'Entrées', 'Sorties', 'URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        ws.row_dimensions[1].height = 30
        
        # Données
        for idx, item in enumerate(self.ark_data, 1):
            row = idx + 1
            
            # Fonction pour convertir en nombre
            def to_num(val, default=0):
                if val == '' or val is None:
                    return default
                try:
                    return int(val)
                except (ValueError, TypeError):
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return default
            
            values = [
                idx,
                item['ark'],
                item['ark_id'],
                item.get('type', ''),
                # Métadonnées OAI-PMH
                item.get('titre', ''),
                item.get('auteur', ''),
                item.get('contributeur', ''),
                item.get('date', ''),
                item.get('editeur', ''),
                item.get('bibliotheque', ''),
                item.get('cote', ''),
                item.get('type_oai', ''),
                item.get('sujet', ''),
                item.get('format_doc', ''),
                item.get('langue', ''),
                item.get('droits', ''),
                item.get('description', ''),
                # Statistiques Matomo (convertis en nombres)
                to_num(item['nb_visits']),
                to_num(item.get('nb_uniq_visitors', '')),
                to_num(item['nb_hits']),
                to_num(item['sum_time_spent']),
                item.get('avg_time_on_page', ''),  # Format texte "00:01:23"
                item.get('bounce_rate', ''),  # Format texte "45 %"
                item.get('exit_rate', ''),  # Format texte "30 %"
                to_num(item.get('entry_nb_visits', '')),
                to_num(item.get('exit_nb_visits', '')),
                item['url']
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                if idx % 2 == 0:
                    cell.fill = alt_fill
                
                # Surligner si titre trouvé
                if col == 5 and value:  # Titre
                    cell.fill = success_fill
                
                if col == 27:  # URL
                    cell.font = link_font
                    if value:
                        cell.hyperlink = value
                elif col in [18, 19, 20, 21, 25, 26]:  # Numériques
                    cell.alignment = Alignment(horizontal='right')
        
        # Largeurs colonnes (27 colonnes maintenant)
        widths = [
            6,   # Rang
            32,  # ARK complet
            28,  # ID ARK
            25,  # Type ressource
            # Métadonnées OAI
            50,  # Titre
            30,  # Auteur
            25,  # Contributeur
            12,  # Date
            35,  # Éditeur
            30,  # Bibliothèque / Source
            25,  # Cote
            30,  # Type document
            40,  # Sujets
            15,  # Format
            10,  # Langue
            25,  # Droits
            50,  # Description
            # Stats Matomo
            10,  # Visites
            16,  # Visiteurs uniques
            12,  # Pages vues
            14,  # Temps total
            12,  # Temps moyen
            12,  # Taux rebond
            12,  # Taux sortie
            10,  # Entrées
            10,  # Sorties
            70   # URL
        ]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        # Filtre et gel
        ws.auto_filter.ref = f"A1:AA{len(self.ark_data)+1}"
        ws.freeze_panes = 'E2'  # Figer les colonnes ARK + scroll sur métadonnées
        
        # === Feuille 2: Résumé ===
        ws2 = wb.create_sheet("Résumé")
        
        ws2['A1'] = "📊 Résumé des statistiques"
        ws2['A1'].font = Font(bold=True, size=16)
        
        ws2['A3'] = "Date d'extraction:"
        ws2['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        ws2['A4'] = "Fichier source:"
        ws2['B4'] = os.path.basename(self.xml_path.get())
        
        ws2['A5'] = "Méthode métadonnées:"
        ws2['B5'] = "API OAI-PMH" if self.scrape_metadata.get() else "Non activée"
        
        ws2['A7'] = "Statistiques globales"
        ws2['A7'].font = Font(bold=True, size=12)
        
        ws2['A8'] = "Nombre de notices ARK:"
        ws2['B8'] = len(self.ark_data)
        
        ws2['A9'] = "Notices avec titre:"
        ws2['B9'] = sum(1 for d in self.ark_data if d.get('titre'))
        
        ws2['A10'] = "Total des visites:"
        ws2['B10'] = sum(d['nb_visits'] for d in self.ark_data)
        
        ws2['A11'] = "Total des pages vues:"
        ws2['B11'] = sum(d['nb_hits'] for d in self.ark_data)
        
        # Par type
        ws2['A13'] = "Par type de ressource"
        ws2['A13'].font = Font(bold=True, size=12)
        
        type_counts = defaultdict(lambda: {'count': 0, 'visits': 0, 'with_title': 0})
        for item in self.ark_data:
            t = item.get('type', 'Autre') or 'Autre'
            type_counts[t]['count'] += 1
            type_counts[t]['visits'] += item['nb_visits']
            if item.get('titre'):
                type_counts[t]['with_title'] += 1
        
        row = 14
        ws2.cell(row=row, column=1, value="Type").font = Font(bold=True)
        ws2.cell(row=row, column=2, value="Notices").font = Font(bold=True)
        ws2.cell(row=row, column=3, value="Avec titre").font = Font(bold=True)
        ws2.cell(row=row, column=4, value="Visites").font = Font(bold=True)
        row += 1
        
        for t, data in sorted(type_counts.items(), key=lambda x: x[1]['visits'], reverse=True):
            ws2.cell(row=row, column=1, value=t)
            ws2.cell(row=row, column=2, value=data['count'])
            ws2.cell(row=row, column=3, value=data['with_title'])
            ws2.cell(row=row, column=4, value=data['visits'])
            row += 1
        
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 15
        
        # === Feuille 3: Top 20 ===
        ws3 = wb.create_sheet("Top 20")
        
        ws3['A1'] = "🏆 Top 20 des ressources les plus consultées"
        ws3['A1'].font = Font(bold=True, size=14)
        
        headers3 = ['Rang', 'Titre / ARK', 'Type', 'Auteur', 'Visites', 'Pages vues']
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='d9e2f3')
        
        for idx, item in enumerate(self.ark_data[:20], 1):
            title = item.get('titre') or item['ark_id']
            ws3.cell(row=idx+3, column=1, value=idx)
            ws3.cell(row=idx+3, column=2, value=title[:60])
            ws3.cell(row=idx+3, column=3, value=item.get('type', '')[:25])
            ws3.cell(row=idx+3, column=4, value=item.get('auteur', '')[:30])
            ws3.cell(row=idx+3, column=5, value=item['nb_visits'])
            ws3.cell(row=idx+3, column=6, value=item['nb_hits'])
        
        ws3.column_dimensions['A'].width = 8
        ws3.column_dimensions['B'].width = 60
        ws3.column_dimensions['C'].width = 28
        ws3.column_dimensions['D'].width = 30
        ws3.column_dimensions['E'].width = 12
        ws3.column_dimensions['F'].width = 12
        
        # === Feuille 4: Composantes BAP/BHP (toujours générée) ===
        if self.components_data:
            ws4 = wb.create_sheet("Composantes")
            
            ws4['A1'] = "📄 Détail par composante (BAP, BHP, pages numérisées)"
            ws4['A1'].font = Font(bold=True, size=14)
            
            ws4['A2'] = f"Total: {len(self.components_data)} composantes"
            ws4['A2'].font = Font(italic=True, color='666666')
            
            headers4 = ['ARK Notice', 'Titre Notice', 'ID Composante', 'Type', 'Visites', 'Visiteurs', 'Pages vues', 'Temps (s)', 'Taux rebond', 'URL']
            for col, h in enumerate(headers4, 1):
                cell = ws4.cell(row=4, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='5b9bd5')
            
            # Trier par visites
            sorted_components = sorted(self.components_data, key=lambda x: x.get('nb_visits', 0), reverse=True)
            
            for idx, comp in enumerate(sorted_components, 1):
                row = idx + 4
                # Déterminer le type de composante
                comp_id = comp.get('component_id', '')
                if comp_id.startswith('BAP'):
                    comp_type = 'Archive (BAP)'
                elif comp_id.startswith('BHP'):
                    comp_type = 'Archive (BHP)'
                elif comp_id.startswith('BMD'):
                    comp_type = 'Archive (BMD)'
                elif comp_id.isdigit() or (len(comp_id) == 4 and comp_id.isdigit()):
                    comp_type = 'Page numérisée'
                else:
                    comp_type = 'Autre'
                
                # Conversion en nombre
                def to_num(val, default=0):
                    if val == '' or val is None:
                        return default
                    try:
                        return int(val)
                    except (ValueError, TypeError):
                        return default
                
                ws4.cell(row=row, column=1, value=comp.get('ark_notice', ''))
                ws4.cell(row=row, column=2, value=comp.get('titre_notice', ''))
                ws4.cell(row=row, column=3, value=comp_id)
                ws4.cell(row=row, column=4, value=comp_type)
                ws4.cell(row=row, column=5, value=to_num(comp.get('nb_visits', 0)))
                ws4.cell(row=row, column=6, value=to_num(comp.get('nb_uniq_visitors', '')))
                ws4.cell(row=row, column=7, value=to_num(comp.get('nb_hits', 0)))
                ws4.cell(row=row, column=8, value=to_num(comp.get('sum_time_spent', 0)))
                ws4.cell(row=row, column=9, value=comp.get('bounce_rate', ''))  # Texte "45 %"
                url_cell = ws4.cell(row=row, column=10, value=comp.get('url', ''))
                if comp.get('url'):
                    url_cell.hyperlink = comp.get('url')
                    url_cell.font = Font(color='0563C1', underline='single')
                
                # Alternance couleurs
                if idx % 2 == 0:
                    for c in range(1, 11):
                        ws4.cell(row=row, column=c).fill = PatternFill('solid', fgColor='deebf7')
            
            ws4.column_dimensions['A'].width = 35
            ws4.column_dimensions['B'].width = 50  # Titre notice
            ws4.column_dimensions['C'].width = 18
            ws4.column_dimensions['D'].width = 18
            ws4.column_dimensions['E'].width = 10
            ws4.column_dimensions['F'].width = 12
            ws4.column_dimensions['G'].width = 12
            ws4.column_dimensions['H'].width = 12
            ws4.column_dimensions['I'].width = 12
            ws4.column_dimensions['J'].width = 75
            
            ws4.auto_filter.ref = f"A4:J{len(self.components_data)+4}"
            ws4.freeze_panes = 'A5'
        
        # Sauvegarder
        wb.save(output_path)
        
        return output_path
    
    def show_preview(self):
        """Affiche un aperçu des données"""
        if not self.xml_path.get():
            messagebox.showwarning("Attention", "Veuillez d'abord sélectionner un fichier XML")
            return
        
        if not self.ark_data:
            # Parser rapidement pour l'aperçu
            try:
                self.ark_data, self.components_data = self.parse_xml(self.xml_path.get())
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de lire le fichier:\n{str(e)}")
                return
        
        # Créer fenêtre d'aperçu
        preview_window = ctk.CTkToplevel(self)
        preview_window.title("Aperçu des données")
        preview_window.geometry("950x550")
        
        # Frame scrollable
        table_frame = ctk.CTkScrollableFrame(preview_window)
        table_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        columns = ['#', 'ARK ID', 'Type', 'Visites', 'Hits', 'Titre']
        for col, header in enumerate(columns):
            ctk.CTkLabel(
                table_frame, 
                text=header,
                font=ctk.CTkFont(weight="bold"),
                width=150 if col > 1 else 50
            ).grid(row=0, column=col, padx=5, pady=8)
        
        # Data rows (max 200)
        max_display = 200
        displayed = min(len(self.ark_data), max_display)
        for row_idx, item in enumerate(self.ark_data[:max_display], 1):
            data_row = [
                row_idx,
                item['ark_id'][:25],
                item.get('type', '')[:20],
                item['nb_visits'],
                item['nb_hits'],
                item.get('titre', '-')[:40] or '-'
            ]
            for col_idx, value in enumerate(data_row):
                ctk.CTkLabel(
                    table_frame,
                    text=str(value),
                    width=150 if col_idx > 1 else 50
                ).grid(row=row_idx, column=col_idx, padx=5, pady=2)
        
        # Footer avec compteur
        if len(self.ark_data) > max_display:
            ctk.CTkLabel(
                preview_window,
                text=f"Affichage limité à {max_display} sur {len(self.ark_data)} notices",
                font=ctk.CTkFont(size=11, slant="italic"),
                text_color="gray"
            ).pack(pady=5)


def main():
    app = MatomoARKExtractor()
    app.mainloop()


if __name__ == '__main__':
    main()
